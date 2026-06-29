import openpyxl
import datetime
import dateutil
import tkinter as tk
from datetime import date
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, colors
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from tkinter import filedialog



location = r'C:\Users\Intern\Documents\Excel_program\Reference_creator\Preference_workbook.xlsx'
state_location = r'C:\Users\Intern\Documents\Excel_program\Reference_creator\statement.xlsx'

wb = Workbook()

root = tk.Tk()
root.withdraw()

firm_name = input("Please enter your firm name:").strip() or "Untitled"
acn = input("Please enter your Australian Company Number (ACN):").strip() or "### ### ###"

day = int(input("Please enter the day of liquidation: "))
month = int(input("Please enter the month of liquidation: "))
year = int(input("Please enter the year of liquidation: "))

liquidate = date(year,month,day)
backdate = date(year,month,day) - relativedelta(months=6)

b_name = []
b_num = []
bb = []
base = 9
register = 0
b_path = []


while register != 'n':
    bank_name = input("Please enter your bank name (supported bank : ANZ, CBA, Bank SA, ING):").strip()  or "Untitled"
    if bank_name.lower() == "commbank" or bank_name.lower() == "commonwealth bank" or  bank_name.lower() == "cba":
        b_name.append('CBA')
        b_path.append(filedialog.askopenfilename())
    elif bank_name.lower() == "australia and new zealand bank" or bank_name.lower() == "australia and new zealand banking group" or  bank_name.lower() == "anz":
        b_name.append('ANZ')
        b_path.append(filedialog.askopenfilename())
    elif bank_name.lower() == "banksa" or bank_name.lower() == "bank sa" or  bank_name.lower() == "bank of south australia":
        b_name.append('Bank SA')
        b_path.append(filedialog.askopenfilename())
    elif bank_name.lower() == "ing" or bank_name.lower() == " internationale nederlanden groep bank":
        b_name.append('ING')
        b_path.append(filedialog.askopenfilename())
    else:
        print('your bank is not supported therefore you will have to manually input the bank statement data')
        b_name.append(bank_name)        
    bank_number = input("Please enter your account number:").strip()  or "########"
    b_num.append(bank_number) 
    bsb = input("Please enter your Bank State Branch (BSB) number:").strip()  or "###-###"
    bb.append(bsb) 
    register = input("do you wish to declare more bank account ( leave blank for yes, type 'n' for no)").strip()

def create_form():

    thick_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))
    counter = 1
    count = 0
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15

    ws['A1'] = firm_name
    ws['A1'].font = Font(name='Arial',bold=True)
    ws['A2'] = "(IN LIQUIDATION)"
    ws['A2'].font = Font(name='Arial',bold=True)
    ws['A3'] = f'ACN {acn}'
    ws['A3'].font = Font(name='Arial',bold=True)
    ws['A5'] = "Preference Review"
    ws['A5'].font = Font(name='Arial',bold=True)
    ws['A6'] = f"Relation back period: {backdate} to {liquidate}"
    ws['A6'].font = Font(name='Arial',bold=True)
    ws['A8'] = "Note: the Company held the following accounts during RBP:"
    
    for i in range(len(b_name)):
        count += 1
        ws[f'A{base + i}'] = f'{count}. {b_name[i]} BSB: {bb[i]} Account number: {b_num[i]}'
        counter += 1

    ws[f'A{base + counter}'] = 'The following summarises all payments during RBP with amounts >= $500 or in round amounts'
    ws[f'A{base + counter}'].font = Font(name='Arial',bold=True) 
    ws[f'A{base + counter + 2}'] = 'Date'
    ws[f'A{base + counter + 2}'].font = Font(name='Arial',bold=True) 
    ws[f'A{base + counter + 2}'].border = thick_border
    ws[f'B{base + counter + 2}'] = 'Description'
    ws[f'B{base + counter + 2}'].font = Font(name='Arial',bold=True)
    ws[f'B{base + counter + 2}'].border = thick_border
    ws[f'C{base + counter + 2}'] = 'Amount ($)'
    ws[f'C{base + counter + 2}'].font = Font(name='Arial',bold=True)
    ws[f'C{base + counter + 2}'].border = thick_border
    ws[f'D{base + counter + 2}'] = 'Account'
    ws[f'D{base + counter + 2}'].font = Font(name='Arial',bold=True)
    ws[f'D{base + counter + 2}'].border = thick_border
    ws[f'E{base + counter + 2}'] = 'Total'
    ws[f'E{base + counter + 2}'].font = Font(name='Arial',bold=True)
    ws[f'E{base + counter + 2}'].border = thick_border
    ws[f'F{base + counter + 2}'] = 'Notes'
    ws[f'F{base + counter + 2}'].font = Font(name='Arial',bold=True)
    ws[f'F{base + counter + 2}'].border = thick_border

    bank_val = base + counter + 3
    return bank_val


def create_sheet():
    global ws
    wb.create_sheet('SORT')
    ws = wb['SORT']
    create_form()
    redFill = PatternFill(patternType='solid', fgColor=colors.Color(rgb='92d050'))

    for i in range(len(b_name)):
        sheetname = f"{b_name[i]} {b_num[i]}"
        wb.create_sheet(sheetname)
        ws = wb[sheetname]
        b_val = create_form()
        ws[f'A{base + i}'].fill = redFill
        ws[f'B{base + i}'].fill = redFill
        if b_name[i] == 'Bank SA':
            bankSA(b_val,i,sheetname)
        if b_name[i] == 'ANZ':
            ANZ(b_val,i,sheetname)
        if b_name[i] == 'CBA':
            CBA(b_val,i,sheetname)
        if b_name[i] == 'ING':
            ING(b_val,i,sheetname)
    
def bankSA(val,num,sheetname):
    bwb = openpyxl.load_workbook(b_path[num])
    bws = bwb.active
    max_rows = bws.max_row
    ws = wb[sheetname]
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    
    counter = 0
    for i in range(2,max_rows+1):
        statement_val = bws[f'C{i}'].value
        if statement_val is None or statement_val == '':
            continue
        statement_val = float(statement_val)
        if  statement_val >= 500 or (statement_val %10==0 and statement_val > 0):
            ws[f'A{val + counter}'] = bws[f'A{i}'].value  
            ws[f'A{val + counter}'].border = thin_border
            ws[f'A{val + counter}'].number_format = 'DD/MM/YYYY' 
            ws[f'B{val + counter}'] = bws[f'B{i}'].value
            ws[f'B{val + counter}'].border = thin_border  
            ws[f'C{val + counter}'] = bws[f'C{i}'].value
            ws[f'C{val + counter}'].border = thin_border
            ws[f'C{val + counter}'].number_format = '$#,##0.00' 
            ws[f'D{val + counter}'] = sheetname
            ws[f'D{val + counter}'].border = thin_border
            ws[f'E{val + counter}'].border = thin_border
            ws[f'F{val + counter}'].border = thin_border    
            counter += 1  

def CBA(val,num,sheetname):
    bwb = openpyxl.load_workbook(b_path[num])
    bws = bwb.active
    max_rows = bws.max_row
    ws = wb[sheetname]
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    
    counter = 0
    for i in range(2,max_rows+1):
        statement_val_str = bws[f'C{i}'].value
        date_val = bws[f'A{i}'].value
        striped = datetime.strptime(date_val,"%d %b %Y").date()
        statement_val = float(statement_val_str.replace("$","").replace(",",""))

        if striped < backdate or striped > liquidate:
            continue
        if statement_val is None or statement_val == '':
            continue
        if  statement_val <= -500 or (statement_val %10==0 and statement_val < 0):
            ws[f'A{val + counter}'] = bws[f'A{i}'].value   
            ws[f'A{val + counter}'].border = thin_border 
            ws[f'A{val + counter}'].number_format = 'DD/MM/YYYY'
            ws[f'B{val + counter}'] = bws[f'B{i}'].value
            ws[f'B{val + counter}'].border = thin_border  
            ws[f'C{val + counter}'] = abs(statement_val)
            ws[f'C{val + counter}'].number_format = '$#,##0.00'
            ws[f'C{val + counter}'].border = thin_border 
            ws[f'D{val + counter}'] = sheetname
            ws[f'D{val + counter}'].border = thin_border
            ws[f'E{val + counter}'].border = thin_border
            ws[f'F{val + counter}'].border = thin_border      
            counter += 1  

def ANZ(val,num,sheetname):
    bwb = openpyxl.load_workbook(b_path[num])
    bws = bwb.active
    max_rows = bws.max_row
    ws = wb[sheetname]
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    
    counter = 0
    for i in range(2,max_rows+1):
        statement_val_str = bws[f'C{i}'].value
        statement_val = float(statement_val_str.replace("$","").replace(",",""))
        if statement_val is None or statement_val == '':
            continue
        if  statement_val <= -500 or (statement_val %10==0 and statement_val < 0):
            ws[f'A{val + counter}'] = bws[f'A{i}'].value   
            ws[f'A{val + counter}'].border = thin_border 
            ws[f'A{val + counter}'].number_format = 'DD/MM/YYYY'
            ws[f'B{val + counter}'] = bws[f'B{i}'].value
            ws[f'B{val + counter}'].border = thin_border  
            ws[f'C{val + counter}'] = abs(statement_val)
            ws[f'C{val + counter}'].number_format = '$#,##0.00'
            ws[f'C{val + counter}'].border = thin_border 
            ws[f'D{val + counter}'] = sheetname
            ws[f'D{val + counter}'].border = thin_border
            ws[f'E{val + counter}'].border = thin_border
            ws[f'F{val + counter}'].border = thin_border      
            counter += 1   
    
def ING(val,num,sheetname):
    bwb = openpyxl.load_workbook(b_path[num])
    bws = bwb.active
    max_rows = bws.max_row
    ws = wb[sheetname]
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    
    counter = 0
    for i in range(2,max_rows+1):
        statement_val_str = bws[f'D{i}'].value
        date_val = bws[f'A{i}'].value
        striped = datetime.strptime(date_val,"%d %m %Y").date()
        statement_val = float(statement_val_str.replace("$","").replace(",",""))

        if striped < backdate or striped > liquidate:
            continue
        if statement_val is None or statement_val == '':
            continue
        if  statement_val <= -500 or (statement_val %10==0 and statement_val < 0):
            ws[f'A{val + counter}'] = bws[f'A{i}'].value   
            ws[f'A{val + counter}'].border = thin_border 
            ws[f'A{val + counter}'].number_format = 'DD/MM/YYYY'
            ws[f'B{val + counter}'] = bws[f'B{i}'].value
            ws[f'B{val + counter}'].border = thin_border  
            ws[f'C{val + counter}'] = abs(statement_val)
            ws[f'C{val + counter}'].number_format = '$#,##0.00'
            ws[f'C{val + counter}'].border = thin_border 
            ws[f'D{val + counter}'] = sheetname
            ws[f'D{val + counter}'].border = thin_border
            ws[f'E{val + counter}'].border = thin_border
            ws[f'F{val + counter}'].border = thin_border      
            counter += 1 

create_sheet()
del wb['Sheet']
wb.save(location)





    