import openpyxl
from openpyxl import load_workbook


location = r'C:\Users\Intern\Documents\Excel_program\comparator\template.xlsx'
Save_location = r'C:\Users\Intern\Documents\Excel_program\comparator\template_output.xlsx'

wb = openpyxl.load_workbook(location)
ws = wb['comparator']

Cdate = []
Camount = []
Cdesc = []

Bdate = []
Bamount = []

counter = 0
cell_base = 2
log = []

compare_val = ws.iter_rows(min_row=2, min_col=1, max_row=855, max_col=3)
base_val = ws.iter_rows(min_row=2, min_col=5, max_row=58, max_col=6)

for a,b,c in compare_val:
    Cdate.append(a.value)
    Camount.append(b.value)
    Cdesc.append(c.value)

for a,b in base_val:
    Bdate.append(a.value)
    Bamount.append(b.value)


for i in range(len(Bdate)):
    for o in range(len(Cdate)): 
        if Cdate[o] == Bdate[i] and Camount[o] == Bamount[i]:
            cellno = cell_base + i
            if 'brodie' in Cdesc[o].lower():
                ws[f"G{cellno}"] = Cdesc[o]
                counter += 1
                break

wb.save(location)
print(f'changed {counter} cell')
